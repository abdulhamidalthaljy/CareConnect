from flask import Blueprint, render_template, request, jsonify, abort, current_app, send_from_directory, url_for, redirect, flash
from flask_login import login_required, current_user
from datetime import datetime
import os
from uuid import uuid4

from src.models.user import Vitals, MedicalFile, PatientProfile, Medicine, Appointment, User
from src.extensions import db
from src.forms import ProfileForm, MedicineForm
from werkzeug.utils import secure_filename
from src.models.user import User
import io
from flask import send_file
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from io import BytesIO as _BytesIO

main = Blueprint('main', __name__)


@main.route('/')
def index():
    return render_template('index.html')


@main.route('/dashboard')
@login_required
def dashboard():
    # If the logged-in user is a doctor, send them to the doctor dashboard
    if (current_user.role or '').strip().lower() == 'doctor':
        return redirect(url_for('doctor.dashboard'))

    # pass current_user and their files to the template so dashboard can display user info and uploads
    files = MedicalFile.query.filter_by(patient_id=current_user.id).order_by(MedicalFile.upload_timestamp.desc()).all()
    # provide empty forms for CSRF tokens and rendering
    from src.forms import ProfileForm, MedicineForm
    profile_form = ProfileForm()
    medicine_form = MedicineForm()
    return render_template('dashboard.html', user=current_user, files=files, form=profile_form, med_form=medicine_form)


@main.route('/add_vital', methods=['POST'])
@login_required
def add_vital():
    # Accept JSON or form-encoded data
    data = request.get_json() or request.form
    v_type = data.get('type')
    value1 = data.get('value1')
    value2 = data.get('value2') or ''

    if not v_type or not value1:
        return jsonify({'error': 'Missing required fields'}), 400
    # server-side validation: numeric values for vitals where appropriate
    # allow value2 to be empty
    try:
        # value1 should be numeric
        float(value1)
        if value2:
            float(value2)
    except ValueError:
        return jsonify({'error': 'Vital values must be numeric'}), 400

    vital = Vitals(patient_id=current_user.id, type=v_type, value1=value1, value2=value2, timestamp=datetime.utcnow())
    db.session.add(vital)
    db.session.commit()

    return jsonify({'status': 'ok', 'id': vital.id, 'timestamp': vital.timestamp.isoformat()})


@main.route('/api/get_vitals')
@login_required
def get_vitals():
    # Optional query parameter `patient_id` for doctors to view other patients
    patient_id = request.args.get('patient_id')
    if patient_id:
        # only doctors can request other patients
        if (current_user.role or '').strip().lower() != 'doctor':
            abort(403)
        pid = int(patient_id)
    else:
        pid = current_user.id

    vitals = Vitals.query.filter_by(patient_id=pid).order_by(Vitals.timestamp.asc()).all()
    data = [
        {
            'id': v.id,
            'type': v.type,
            'value1': v.value1,
            'value2': v.value2,
            'timestamp': v.timestamp.isoformat()
        }
        for v in vitals
    ]
    return jsonify(data)


# ---------------- File upload routes ----------------
ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png'}

# Basic mapping of extensions to accepted MIME types to cross-check browser-provided mimetype
EXT_TO_MIMETYPE = {
    '.pdf': ['application/pdf'],
    '.jpg': ['image/jpeg'],
    '.jpeg': ['image/jpeg'],
    '.png': ['image/png'],
}


def allowed_file(filename, mimetype):
    ext = os.path.splitext(filename.lower())[1]
    if ext not in ALLOWED_EXTENSIONS:
        return False
    # if mimetype is provided, ensure it roughly matches extension
    if mimetype:
        allowed = EXT_TO_MIMETYPE.get(ext)
        if allowed and mimetype.split(';')[0] not in allowed:
            return False
    return True


@main.route('/upload_file', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    # validate extension and mimetype
    if not allowed_file(file.filename, file.mimetype):
        return jsonify({'error': 'File type not allowed'}), 400

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1]
    storage_name = f"{uuid4().hex}{ext}"

    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    # store per-user to reduce accidental collisions and help cleanup
    user_folder = os.path.join(upload_folder, f'user_{current_user.id}')
    os.makedirs(user_folder, exist_ok=True)
    save_path = os.path.join(user_folder, storage_name)
    # extra safeguard: do not allow path traversal via secure_filename
    if not os.path.commonpath([os.path.abspath(user_folder)]) == os.path.commonpath([os.path.abspath(user_folder), os.path.abspath(save_path)]):
        return jsonify({'error': 'Invalid file path'}), 400
    # Validate file content based on extension. Import Pillow at runtime to avoid hard dependency at module import.
    ext_l = ext.lower()
    try:
        data = file.read()
    except Exception:
        return jsonify({'error': 'Failed reading file'}), 400

    # Simple PDF check: file should start with %PDF
    if ext_l == '.pdf':
        if not data.startswith(b'%PDF'):
            return jsonify({'error': 'Invalid PDF file'}), 400

    # For images, try to validate with Pillow if available
    if ext_l in ('.jpg', '.jpeg', '.png'):
        try:
            from PIL import Image as _Image
        except Exception:
            _Image = None
        if _Image:
            try:
                img = _Image.open(_BytesIO(data))
                img.verify()
            except Exception:
                return jsonify({'error': 'Invalid image file'}), 400

    # Persist bytes to disk
    try:
        with open(save_path, 'wb') as fh:
            fh.write(data)
    except Exception:
        return jsonify({'error': 'Failed to save file'}), 500

    mf = MedicalFile(patient_id=current_user.id, original_filename=filename, storage_filename=storage_name)
    db.session.add(mf)
    db.session.commit()

    return jsonify({'status': 'ok', 'file_id': mf.id, 'original_filename': mf.original_filename})


@main.route('/download_file/<int:file_id>')
@login_required
def download_file(file_id):
    mf = MedicalFile.query.get_or_404(file_id)
    # allow owner or doctor to download
    if mf.patient_id != current_user.id and (current_user.role or '').strip().lower() != 'doctor':
        abort(403)

    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    return send_from_directory(upload_folder, mf.storage_filename, as_attachment=True, attachment_filename=mf.original_filename)


@main.route('/update_profile', methods=['GET', 'POST'])
@login_required
def update_profile():
    form = ProfileForm()
    if form.validate_on_submit():
        full_name = form.full_name.data.strip()
        address = form.address.data.strip()
        allergies = form.allergies.data.strip() if form.allergies.data else ''
        health_history = form.health_history.data.strip() if form.health_history.data else ''

        profile = PatientProfile.query.filter_by(user_id=current_user.id).first()
        if not profile:
            profile = PatientProfile(user_id=current_user.id, full_name=full_name, address=address, allergies=allergies, health_history=health_history)
            db.session.add(profile)
        else:
            profile.full_name = full_name
            profile.address = address
            profile.allergies = allergies
            profile.health_history = health_history

        db.session.commit()
        flash('Profile updated successfully.', 'success')
        return redirect(url_for('main.dashboard'))

    # on GET, prefill form from existing profile
    profile = PatientProfile.query.filter_by(user_id=current_user.id).first()
    if profile:
        form.full_name.data = profile.full_name
        form.address.data = profile.address
        form.allergies.data = profile.allergies
        form.health_history.data = profile.health_history
    return render_template('dashboard.html', user=current_user, files=MedicalFile.query.filter_by(patient_id=current_user.id).order_by(MedicalFile.upload_timestamp.desc()).all(), form=form)


@main.route('/add_medicine', methods=['POST'])
@login_required
def add_medicine():
    # Only patients add medicines to their profile
    if (current_user.role or '').strip().lower() != 'patient':
        abort(403)
    form = MedicineForm()
    if form.validate_on_submit():
        name = form.name.data.strip()
        dosage = form.dosage.data.strip() if form.dosage.data else ''
        med = Medicine(patient_id=current_user.id, name=name, dosage=dosage)
        db.session.add(med)
        db.session.commit()
        flash('Medicine added.', 'success')
        return redirect(url_for('main.dashboard'))
    flash('Invalid medicine data.', 'danger')
    return redirect(url_for('main.dashboard'))


@main.route('/delete_medicine/<int:med_id>', methods=['POST'])
@login_required
def delete_medicine(med_id):
    med = Medicine.query.get_or_404(med_id)
    # Only owner patient or a doctor may delete
    if med.patient_id != current_user.id and (current_user.role or '').strip().lower() != 'doctor':
        abort(403)
    db.session.delete(med)
    db.session.commit()
    flash('Medicine removed.', 'info')
    return redirect(url_for('main.dashboard'))


@main.route('/export_excel')
@login_required
def export_excel():
    # allow doctors to export for a given patient via ?patient_id=
    patient_id = request.args.get('patient_id')
    if patient_id:
        if (current_user.role or '').strip().lower() != 'doctor':
            abort(403)
        pid = int(patient_id)
        # verify patient exists and is actually a patient
        patient = User.query.get(pid)
        if not patient or (patient.role or '').strip().lower() != 'patient':
            abort(404)
    else:
        pid = current_user.id

    profile = PatientProfile.query.filter_by(user_id=pid).first()
    medicines = Medicine.query.filter_by(patient_id=pid).all()
    vitals = Vitals.query.filter_by(patient_id=pid).order_by(Vitals.timestamp.asc()).all()

    wb = Workbook()
    # Profile sheet
    ps = wb.active
    ps.title = 'Profile'
    ps.append(['Field', 'Value'])
    user_obj = User.query.get(pid)
    ps.append(['Username', user_obj.username if user_obj else ''])
    ps.append(['Full name', profile.full_name if profile else ''])
    ps.append(['Address', profile.address if profile else ''])
    ps.append(['Allergies', profile.allergies if profile else ''])
    ps.append(['Health history', profile.health_history if profile else ''])

    # Medicines sheet
    ms = wb.create_sheet('Medicines')
    ms.append(['Name', 'Dosage'])
    for m in medicines:
        ms.append([m.name, m.dosage])

    # Vitals sheet
    vs = wb.create_sheet('Vitals')
    vs.append(['Type', 'Value1', 'Value2', 'Timestamp'])
    for v in vitals:
        vs.append([v.type, v.value1, v.value2 or '', v.timestamp.isoformat()])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f'careconnect_patient_{pid}.xlsx'
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main.route('/export_pdf')
@login_required
def export_pdf():
    # allow doctors to export for a given patient via ?patient_id=
    patient_id = request.args.get('patient_id')
    if patient_id:
        if (current_user.role or '').strip().lower() != 'doctor':
            abort(403)
        pid = int(patient_id)
        # verify patient exists and is actually a patient
        patient = User.query.get(pid)
        if not patient or (patient.role or '').strip().lower() != 'patient':
            abort(404)
    else:
        pid = current_user.id

    profile = PatientProfile.query.filter_by(user_id=pid).first()
    vitals = Vitals.query.filter_by(patient_id=pid).order_by(Vitals.timestamp.asc()).all()

    bio = io.BytesIO()
    p = canvas.Canvas(bio)
    user_obj = User.query.get(pid)
    title = f'Vitals Report for {user_obj.username if user_obj else pid}'
    p.setFont('Helvetica-Bold', 14)
    p.drawString(40, 800, title)
    p.setFont('Helvetica', 10)
    y = 780
    if profile:
        p.drawString(40, y, f'Full name: {profile.full_name or ""}'); y -= 16
        p.drawString(40, y, f'Address: {profile.address or ""}'); y -= 16
        p.drawString(40, y, f'Allergies: {profile.allergies or ""}'); y -= 20
    else:
        p.drawString(40, y, 'No profile information'); y -= 20

    p.drawString(40, y, 'Vitals:'); y -= 16
    if vitals:
        for v in vitals:
            line = f"{v.timestamp.strftime('%Y-%m-%d %H:%M')} - {v.type} - {v.value1}{('/'+v.value2) if v.value2 else ''}"
            p.drawString(48, y, line)
            y -= 14
            if y < 60:
                p.showPage()
                p.setFont('Helvetica', 10)
                y = 800
    else:
        p.drawString(48, y, 'No vitals recorded')

    p.showPage()
    p.save()
    bio.seek(0)

    filename = f'careconnect_vitals_{pid}.pdf'
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/pdf')